import openpyxl
"""
https://openpyxl.readthedocs.io/en/stable/
https://www.pythonexcel.com/openpyxl.php
"""

o = openpyxl.load_workbook("./TestData.xlsx")

o.create_sheet("Demo",0)
vk  = o.sheetnames
print(vk)

sh1=o["URLs"]
sh1['A2']="www.amazon.in"
o.save("./TestData.xlsx")


#
# sh=o["Home"]
#
# val =sh['A2'].value
#
#
#
# rows = sh.max_row
# columns = sh.max_column
#
# print(rows)
# print(columns)




#
# # object of the sheet
# sh =o["Login"]
# print(sh.title)



